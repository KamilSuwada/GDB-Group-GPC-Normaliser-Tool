from openpyxl import Workbook, load_workbook
import os, uuid


class XLSXFile():

    def __init__(self, path: str) -> None:
        self.id = uuid.uuid4()
        self.path = path
        self.results = []
        self.wb = self.make_workbook(path=self.path)
        self.data = self.read_workbook(wb=self.wb)
        
    

    def make_workbook(self, path: str) -> Workbook:
        try:
            wb = load_workbook(filename=path)
            return wb
        except:
            print(f"File at path:\n\n{path}\n\ncould not have been read! Please make sure to select only .xlsx files!")
            exit()


    def read_workbook(self, wb: Workbook) -> dict:
        result = {
            "path": self.path
        }

        sh = wb["Raw Data"]

        column = 1
        row = 1
        catchRow = 0
        max_index = 5000

        time = []
        channels = []
        data_points = {}

        while True:
            value = sh.cell(column=column, row=row).value

            if value == "Channel ID":
                catchRow = row
                while True:
                    row += 1
                    val = sh.cell(column=column, row=row).value

                    if val == None:
                        row = catchRow
                        break

                    channel = {
                        "id": sh.cell(column=column, row=row).value,
                        "type": sh.cell(column=column+1, row=row).value
                    }

                    channels.append(channel)

            if value == "RT (mins)":
                catchRow = row

                while True: # Time extraction...
                    row += 1
                    val = sh.cell(column=column, row=row).value

                    if val == None:
                        max_index = row
                        row = catchRow
                        break

                    time.append(float(val))

                i = 1
                
                for channel in channels:
                    i += 1
                    data = []

                    while True: # Traces extraction
                        row += 1
                        val = sh.cell(column=i, row=row).value

                        if val == None:
                            row = catchRow
                            break

                        data.append(float(val))

                    key = str(channel["type"])
                    data_points[key] = data

            
            row += 1

            if row > max_index:
                break


        result["time"] = time
        result["channels"] = channels
        result["data_points"] = data_points
        print(f"Successful data extraction from: {os.path.basename(self.path)}")
        return result



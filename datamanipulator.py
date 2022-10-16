from xlsxfile import XLSXFile
from openpyxl import Workbook
from openpyxl.chart import Reference, LineChart
from openpyxl.utils import get_column_letter
import os


class DataManipulator():

    def __init__(self) -> None:
        self.kinetic_keys = []
        self.detector_keys = []
        self.kinetic_results = []
        self.files_data = []
        self.max_areas = {}
        self.max_heights = {}
        self.number_of_files = -1


    def clear_all_kinetics(self) -> None:
        self.files_data = []
        self.max_areas = {}
        self.max_heights = {}


##################################### ARGS INPUT CHECKS:


    def check_mode(self, arg: str):
        mode = arg.lower().strip()
        if mode == "height" or mode == "kinetic" or mode == "both":
            return mode
        else:
            exit(f"Incorrect mode has been provided: {mode}. Please adjust you input to:\n\nkineic\nheight\nboth")


    def check_ranges_input(self, args, combination: bool) -> list:
        if combination:
            input_str = " ".join(args)
            components = input_str.split("!")
            start_times = components[0].split()
            stop_times = components[1].split()
            if len(start_times) == 0 or len(stop_times) == 0:
                exit("It seems that you have not provided any of either the stop or start times. Please refer to the help.")
            ranges_combinations = self.make_start_stop_times_combinations(starts=start_times, stops=stop_times)
            return ranges_combinations
        else:
            if len(args) % 2 != 0:
                exit("Invalid number or start and stop times was provided. Every start time need a stop time right after it. Please refer to help.")
            output = []
            number_of_ranges = int(len(args) / 2)
            for i in range(0, number_of_ranges):
                try:
                    start = float(args[i + i])
                except ValueError:
                    exit(f"Could not convert {args[i + i]} to a floating point number... Double check your input.")

                try:
                    stop = float(args[i + i + 1])
                except ValueError:
                    exit(f"Could not convert {args[i + i + 1]} to a floating point number... Double check your input.")

                if start >= stop:
                    exit(f"Start time cannot be greater than stop time: {start} and {stop}")
                output.append((start, stop))
            
            return output

    
    def make_start_stop_times_combinations(self, starts: list, stops: list) -> list:
        output = []

        for start in starts:
            for stop in stops:
                try:
                    start = float(start)
                except ValueError:
                    exit(f"Could not convert {start} to a floating point number... Double check your input.")

                try:
                    stop = float(stop)
                except ValueError:
                    exit(f"Could not convert {stop} to a floating point number... Double check your input.")
                
                range = (start, stop)
                if range[0] > range[1]:
                    exit(f"Start time cannot be greater than stop time: {range[0]} and {range[1]}")
                output.append(range)
        
        return output
                
                



##################################### DATA EXTRACTION METHODS:

    def extract_data(self, file: XLSXFile, start: float, stop: float) -> dict:
        time = file.data["time"]
        extracted_time = self.extract_time_data(data=time, start=start, stop=stop)
        indexes = self.find_start_and_stop_indexes(data=time, start_value=extracted_time[0], stop_value=extracted_time[len(extracted_time) - 1])

        output = {
            "start_time": start,
            "stop_time": stop,
            "time": extracted_time,
        }

        for d in file.data["channels"]:
            key = d["type"]
            detector_data = file.data["data_points"][key]
            extracted_data = self.extract_from_data_on_indexes(data=detector_data, start_index=indexes[0], stop_index=indexes[1])
            output[key] = extracted_data
        
        return output


    ### Data extraction method.
    def extract_time_data(self, data: list, start: float, stop: float) -> list:
        return [point for point in data if point >= start if point <= stop]


    ### Method returs extracted data from a list based on specified indexes.
    def extract_from_data_on_indexes(self, data: list, start_index: int, stop_index: int) -> list:
        if stop_index < start_index or len(data) - 1 < stop_index or start_index < 0:
            print(f"Fundamentally wrong indexes were passed to extract data, are you sure {start_index} and {stop_index} are correct?")
            exit()

        return data[start_index : stop_index + 1]
    

    ### Finds start and stop indexes of the extracted data using extract_time_data method.
    def find_start_and_stop_indexes(self, data: list, start_value: float, stop_value: float) -> tuple:
        start_index = -1
        stop_index = -1
        index = 0

        for point in data:
            if point == start_value:
                start_index = index
            if point == stop_value:
                stop_index = index
                break
            
            index += 1
        
        return (start_index, stop_index)





##################################### Height normalisation:


    def height_normalise(self, file: XLSXFile, start: float, stop: float):
        extracted_raw = self.extract_data(file=file, start=start, stop=stop)
        for d in file.data["channels"]:
            key = d["type"]
            h_normalised_data = self.do_height_normalisation(data=extracted_raw[key])
            result = {
                "mode": "height",
                "detector_type": key,
                "start_time": extracted_raw["start_time"],
                "stop_time": extracted_raw["stop_time"],
                "time": extracted_raw["time"],
                "height_normalised": h_normalised_data
            }
            file.results.append(result)


    def do_height_normalisation(self, data: list) -> list:
        min_shifted = self.shift_minimum_to_zero(data=data)
        return self.height_normalise_set(data=min_shifted)
    

    def height_normalise_set(self, data: list) -> list:
        max = self.find_max(data=data)
        return [point / max for point in data] # by dividing all points by max value of the set, the result values will be between 0 and 1 for min-shifted data.


    def shift_minimum_to_zero(self, data: list) -> list:
        min = self.find_min(data=data)
        return [point - min for point in data] # by substarcting minimum value of the set from all points, the actual minimum will shift to 0.


    def find_min(self, data: list) -> float:
        return min(data)


    def find_max(self, data: list) -> float:
        return max(data)





##################################### Kinetic normalisation:


    def kinetic_normalise(self, files: list, start: float, stop: float):
        kinetic_key = f"{start} - {stop}"
        self.kinetic_keys.append(kinetic_key)
        time = []

        for file in files:
            dT = file.data["time"][1] - file.data["time"][0]
            extracted_raw = self.extract_data(file=file, start=start, stop=stop)
            time = extracted_raw["time"]
            for d in file.data["channels"]:
                key = d["type"]
                if key not in self.detector_keys:
                    self.detector_keys.append(key)
                min_shifted = self.shift_minimum_to_zero(data=extracted_raw[key])
                area = self.compute_area(data=min_shifted, dT=dT)

                try:
                    previous_max_area = self.max_areas[key]
                    if previous_max_area < area:
                        self.max_areas[key] = area
                except:
                    self.max_areas[key] = area

                chunk_data = {
                    "file_name": os.path.basename(file.path),
                    "file_id": str(file.id),
                    "channel": key,
                    "min_shifted": min_shifted,
                    "area": area,
                    "area_normalised": [],
                    "height_normalised": [],
                }
                
                self.files_data.append(chunk_data)

        for data in self.files_data:
            key = data["channel"]
            coeff = self.max_areas[key] / data["area"]
            data["area_normalised"] = [point * coeff for point in data["min_shifted"]]

        for data in self.files_data:
            key = data["channel"]
            height = self.find_max(data=data["area_normalised"])

            try:
                previous_max_height = self.max_heights[key]
                if previous_max_height < height:
                    self.max_heights[key] = height
            except:
                self.max_heights[key] = height

        for data in self.files_data:
            key = data["channel"]
            max_height = self.max_heights[key]
            data["height_normalised"] = [point / max_height for point in data["area_normalised"]]

            result = {
                "file_name": data["file_name"],
                "mode": "kinetic",
                "time_key": kinetic_key,
                "detector_type": key,
                "start_time": start,
                "stop_time": stop,
                "time": time,
                "kinetics_normalised": data["height_normalised"]
            }

            self.kinetic_results.append(result)

        self.clear_all_kinetics()

    
    def compute_area(self, data: list, dT: float) -> float:
        previous_point = 0
        area = 0

        for point in data:
            area_chunk = (previous_point + point) * dT * 0.5
            area += area_chunk
            previous_point = point
        
        return area





##################################### Saving results:


    def save_height_data_to_file(self, save_directory: str, files: list):
        wb = Workbook()
        
        for file in files:
            sheet_name = os.path.basename(file.path)
            ws = wb.create_sheet()
            ws.title = sheet_name
            self.write_height_data_set_to_worksheet(ws=ws, file=file)

        
        filename = f"height_results.xlsx"
        wb_path = os.path.join(save_directory, filename)
        sheet = wb["Sheet"]
        wb.remove_sheet(sheet)
        wb.save(filename=wb_path)
        print(f"Saved height results to: {wb_path}")


    def write_height_data_set_to_worksheet(self, ws, file: XLSXFile):
        column = 1
        row = 1
        count = 0

        for result in file.results:
            column += count * 10 + count

            ws.cell(column=column, row=row).value = "RT (minutes)"
            ws.cell(column=column + 1, row=row).value = str(result["detector_type"])
            ws.cell(column=column + 2, row=row).value = "Parameters"
            ws.cell(column=column + 3, row=row).value = "Values"

            ws.cell(column=column + 2, row=row + 1).value = "Mode"
            ws.cell(column=column + 3, row=row + 1).value = str(result["mode"])

            ws.cell(column=column + 2, row=row + 2).value = "Detector type"
            ws.cell(column=column + 3, row=row + 2).value = str(result["detector_type"])

            ws.cell(column=column + 2, row=row + 3).value = "Start time"
            ws.cell(column=column + 3, row=row + 3).value = result["start_time"]

            ws.cell(column=column + 2, row=row + 4).value = "Stop time"
            ws.cell(column=column + 3, row=row + 4).value = result["stop_time"]

            ws.cell(column=column + 2, row=row + 5).value = "Number of points"
            ws.cell(column=column + 3, row=row + 5).value = len(result["time"])

            index = 0
            row = 2

            for time_point in result["time"]:
                ws.cell(column=column, row=row + index).value = float(time_point)
                ws.cell(column=column + 1, row=row + index).value = float(result["height_normalised"][index])
                index += 1

            values = Reference(ws, min_col=column + 1, min_row=1, max_col=column + 1, max_row=row + index)
            x_values = Reference(ws, min_col=column, min_row=2, max_col=column, max_row=row + index)

            chart = LineChart()
            chart.add_data(values, titles_from_data = True)
            chart.set_categories(x_values)
            chart.title = f"{result['detector_type']}: {result['start_time']} - {result['stop_time']}"
            chart.y_axis.title = f"{result['detector_type']} (normalised)"
            chart.x_axis.title = "Retention Time (minutes)"
            chart.legend = None
            chart.y_axis.scaling.min = 0
            chart.y_axis.scaling.max = 1.01
            chart.height = 15
            chart.y_axis.majorUnit = 1
            number_of_points = len(result["time"]) / (float(result["stop_time"]) - float(result["start_time"]))
            chart.x_axis.tickLblSkip = int(number_of_points)
            placement = get_column_letter(column)
            for series in chart.series:
                series.graphicalProperties.line.width = 1
            ws.add_chart(chart, f"{placement}7")

            row = 1
            column = 1
            count += 1

        
    def save_kinetics_data_to_file(self, save_directory: str):
        wb = Workbook()
        
        for key in self.detector_keys:
            sheet_name = str(key)
            ws = wb.create_sheet()
            ws.title = sheet_name
            self.write_kinetic_data_set_to_worksheet(ws=ws, detector_key=key)

        
        filename = f"kinetic_results.xlsx"
        wb_path = os.path.join(save_directory, filename)
        sheet = wb["Sheet"]
        wb.remove_sheet(sheet)
        wb.save(filename=wb_path)
        print(f"Saved kinetic results to: {wb_path}")


    def write_kinetic_data_set_to_worksheet(self, ws, detector_key: str):
        column = 1
        row = 1
        index = 2
        count = 0

        for key in self.kinetic_keys:
            column += count * (self.number_of_files + 2) + count
            number_of_points = 0
            ws.cell(column=column, row=row).value = "Time range"
            ws.cell(column=column, row=row + 1).value = str(key)
            ws.cell(column=column + 1, row=row).value = "Time (minutes)"

            for result in self.kinetic_results:
                if result["detector_type"] == detector_key:
                    if result["time_key"] == key:
                        time = result["time"]
                        i = 1
                        for point in time:
                            ws.cell(column=column + 1, row=row + i).value = float(point)
                            i += 1
                        number_of_points = float(len(time)) / (float(result["stop_time"]) - float(result["start_time"]))
                        break

            x_values = Reference(ws, min_col=column + 1, min_row=2, max_col=column + 1, max_row=row + 1)


            for result in self.kinetic_results:
                if result["detector_type"] == detector_key:
                    if result["time_key"] == key:
                        ws.cell(column=column + index, row=row).value = str(result["file_name"])
                        data = result["kinetics_normalised"]
                        i = 1
                        for point in data:
                            ws.cell(column=column + index, row=row + i).value = float(point)
                            i += 1
                        index += 1

            values = Reference(ws, min_col=column + 2, min_row=1, max_col=column + index - 1, max_row=row + i)

            chart = LineChart()
            chart.add_data(values, titles_from_data = True)
            chart.set_categories(x_values)
            chart.title = f"{detector_key}: {key}"
            chart.y_axis.title = f"{detector_key} (normalised)"
            chart.x_axis.title = "Retention Time (minutes)"
            chart.legend.position = "b"
            chart.y_axis.scaling.min = 0
            chart.y_axis.scaling.max = 1.01
            chart.height = 20
            chart.width = 30
            chart.y_axis.majorUnit = 1
            chart.x_axis.tickLblSkip = number_of_points
            placement = get_column_letter(column)
            for series in chart.series:
                series.graphicalProperties.line.width = 1
            ws.add_chart(chart, f"{placement}7")

            index = 2
            column = 1
            row = 1
            count += 1




        
    
